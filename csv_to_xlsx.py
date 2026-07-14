from __future__ import annotations

import argparse
import csv
import json
import time
import zipfile
from concurrent.futures import ProcessPoolExecutor, as_completed
from pathlib import Path

import openpyxl
import xlsxwriter


EXCEL_MAX_ROWS = 1_048_576

NUMERIC_HEADERS = {
    "本位币金额", "交易币金额", "借方本位币", "贷方本位币", "借方交易币", "贷方交易币",
    "期初余额", "期末余额", "借方发生额", "贷方发生额", "期间净发生额", "净发生额",
    "本年借方累计", "本年贷方累计", "本年累计发生额", "差异", "ACDOCA余额结转",
}
INTEGER_HEADERS = {"记录数", "BCF记录数"}
LONG_TEXT_HEADERS = {
    "凭证抬头文本", "行项目文本", "科目名称", "公司名称", "事务码描述", "凭证类型描述",
    "过账码描述", "供应商名称", "供应商名称1", "供应商名称2", "供应商名称3", "供应商名称4",
    "客户名称", "物料名称", "资产名称", "用户姓名", "分类账名称",
}


def parse_number(value: str) -> float | None:
    if value == "":
        return None
    try:
        return float(value.replace(",", ""))
    except ValueError:
        return None


def column_width(header: str) -> int:
    if header in LONG_TEXT_HEADERS:
        return 28
    if header in NUMERIC_HEADERS:
        return 16
    if "日期" in header or "时间" in header:
        return 12
    if any(token in header for token in ("编号", "科目", "订单", "凭证", "代码", "用户", "资产")):
        return 15
    return max(10, min(18, len(header) * 2 + 2))


def convert_one(source: Path, destination: Path) -> dict[str, object]:
    started = time.monotonic()
    destination.parent.mkdir(parents=True, exist_ok=True)
    temp_path = destination.with_suffix(".xlsx.tmp")
    if temp_path.exists():
        temp_path.unlink()

    workbook = xlsxwriter.Workbook(
        temp_path,
        {"constant_memory": True, "strings_to_numbers": False, "strings_to_urls": False},
    )
    workbook.use_zip64()
    sheet_name = "序时账" if "序时账_" in source.name else "数据"
    worksheet = workbook.add_worksheet(sheet_name)
    worksheet.freeze_panes(1, 0)
    worksheet.set_zoom(80)

    header_format = workbook.add_format({
        "bold": True, "font_color": "#FFFFFF", "bg_color": "#1F4E78",
        "align": "center", "valign": "vcenter", "text_wrap": True,
        "border": 1, "border_color": "#B4C6E7",
    })
    money_format = workbook.add_format({"num_format": "#,##0.00;[Red]-#,##0.00", "align": "right"})
    integer_format = workbook.add_format({"num_format": "#,##0", "align": "right"})
    text_format = workbook.add_format({"num_format": "@"})

    rows_written = 0
    with source.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.reader(handle)
        header = next(reader)
        if len(header) > 16_384:
            raise ValueError(f"列数超过 Excel 上限: {source}")
        worksheet.set_row(0, 30)
        worksheet.write_row(0, 0, header, header_format)
        numeric_indexes = {i for i, value in enumerate(header) if value in NUMERIC_HEADERS}
        integer_indexes = {i for i, value in enumerate(header) if value in INTEGER_HEADERS}
        for index, value in enumerate(header):
            worksheet.set_column(index, index, column_width(value), text_format)

        for excel_row, row in enumerate(reader, start=1):
            if excel_row >= EXCEL_MAX_ROWS:
                raise ValueError(f"数据超过 Excel 行数上限: {source}")
            if len(row) != len(header):
                raise ValueError(f"CSV 列数异常: {source} 第 {excel_row + 1} 行")
            for col, value in enumerate(row):
                if col in numeric_indexes:
                    number = parse_number(value)
                    if number is None:
                        worksheet.write_blank(excel_row, col, None, money_format) if value == "" else worksheet.write_string(excel_row, col, value)
                    else:
                        worksheet.write_number(excel_row, col, number, money_format)
                elif col in integer_indexes:
                    number = parse_number(value)
                    if number is None:
                        worksheet.write_blank(excel_row, col, None, integer_format) if value == "" else worksheet.write_string(excel_row, col, value)
                    else:
                        worksheet.write_number(excel_row, col, int(number), integer_format)
                else:
                    worksheet.write_string(excel_row, col, value, text_format)
            rows_written = excel_row
            if rows_written % 100_000 == 0:
                print(json.dumps({"file": source.name, "rows_written": rows_written}, ensure_ascii=False), flush=True)

    worksheet.autofilter(0, 0, rows_written, len(header) - 1)
    workbook.close()
    temp_path.replace(destination)

    # Independent structural verification: ZIP integrity and read-only workbook dimensions/header.
    with zipfile.ZipFile(destination) as archive:
        bad_member = archive.testzip()
        if bad_member:
            raise ValueError(f"XLSX ZIP 损坏: {destination} -> {bad_member}")
    check = openpyxl.load_workbook(destination, read_only=True, data_only=False)
    sheet = check[check.sheetnames[0]]
    checked_header = [cell.value if cell.value is not None else "" for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    result = {
        "source": str(source), "output": str(destination), "data_rows": rows_written,
        "excel_rows": sheet.max_row, "columns": sheet.max_column,
        "header_match": checked_header == header, "seconds": round(time.monotonic() - started, 1),
    }
    check.close()
    if result["excel_rows"] != rows_written + 1 or result["columns"] != len(header) or not result["header_match"]:
        raise ValueError(f"XLSX 验证失败: {result}")
    return result


def main() -> None:
    parser = argparse.ArgumentParser(description="将 S/4HANA 序时账 CSV 流式并行转换为 Excel")
    parser.add_argument("--input-dir", required=True)
    parser.add_argument("--output-dir", required=True)
    parser.add_argument("--workers", type=int, default=1)
    parser.add_argument("--overwrite", action="store_true")
    args = parser.parse_args()
    input_dir = Path(args.input_dir).resolve()
    output_dir = Path(args.output_dir).resolve()
    sources = sorted(input_dir.rglob("*.csv"))
    if not sources:
        raise FileNotFoundError(f"没有找到 CSV: {input_dir}")
    jobs: list[tuple[Path, Path]] = []
    manifest = []
    for source in sources:
        destination = output_dir / source.relative_to(input_dir).with_suffix(".xlsx")
        if destination.exists() and not args.overwrite:
            print(f"[跳过已完成] {destination.name}", flush=True)
            continue
        jobs.append((source, destination))

    if args.workers <= 1:
        for current, (source, destination) in enumerate(jobs, 1):
            print(f"[{current}/{len(jobs)}] {source.name}", flush=True)
            result = convert_one(source, destination)
            manifest.append(result)
            print(json.dumps(result, ensure_ascii=False), flush=True)
    else:
        print(f"并行进程: {args.workers}，待转换: {len(jobs)}", flush=True)
        with ProcessPoolExecutor(max_workers=args.workers) as executor:
            futures = {executor.submit(convert_one, source, destination): source for source, destination in jobs}
            for completed, future in enumerate(as_completed(futures), 1):
                source = futures[future]
                result = future.result()
                manifest.append(result)
                print(f"[{completed}/{len(jobs)} 完成] {source.name}", flush=True)
                print(json.dumps(result, ensure_ascii=False), flush=True)
    manifest_path = output_dir / "Excel转换验证清单.json"
    manifest_path.write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps({"completed": len(manifest), "manifest": str(manifest_path)}, ensure_ascii=False), flush=True)


if __name__ == "__main__":
    main()
